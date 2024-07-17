# -*- coding: utf-8 -*-
from datetime import datetime, timedelta # datetime:處理日期及時間; timedelta: 計算二個日期或時間的差異
import twstock     # 台灣股市 Python 庫
import talib       # 技術分析 Python 庫
import numpy as np # 科學計算庫
import yfinance as yf
import os
import pygsheets              # Google Sheets操作
import pandas as pd           # 數據分析和處理庫
import requests               # HTTP 請求庫
from bs4 import BeautifulSoup # HTML 和 XML 解析庫
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, PatternFill, Border, Side

"""
class StockUtils: 取得「指定股票」的各項「技術指標」
    adjustDate()              : 調整日期，如傳入日期為星期六/日，調整為「星期五」
    getStartDate()            : 取得「開始日期」及「抓取資料的開始日期」
    getStockAnalyze()         : 取得「指定股票」的基本資料 及 各項技術指標
    countIncreaseDaysHistory(): 針對傳入的list，每一天都取得「多頭/空頭數量」連續增加天數，回傳 dict
    countIncreaseDays()       : 只取得「最新日期」的「多頭/空頭數量」連續增加天數，回傳 int
    countIndexDays()          : 取得「多頭/空頭數量」連續 >=index 的天數
    saveToExcel()             : 將數據保存到 Excel 文件中
    saveToGoogle()            : 將數據保存到 Google Sheet 文件中
    getStockData4Yahoo()      : 爬Yahoo網頁,取得當天股票資訊

"""

class StockUtils:
    def __init__(self):
        """
        取得「指定股票」的各項「技術指標」.

        Attributes:
            sHr1 (str): 分隔線 in printed output.
            sHr2 (str): 分隔線 in printed output.
            dToday (datetime.datetime): 當天日期
            sToday (str): 當天日期

        Created on: 2023-06-23
        Last modified: 2024-06-23
        """
        # print('\t ===== class StockUtils() ===== ')
        # get 當天日期
        self.dToday = datetime.today() #<class 'datetime.datetime'>
        self.sToday = self.dToday.strftime('%Y-%m-%d') #<class 'str'>
        # print(F"\t dToday = ({type(self.dToday)}) {self.dToday}")
        # print(F"\t sToday = ({type(self.sToday)}) {self.sToday}")

        # 排版相關變數
        self.sHr1 = "==================================================";
        self.sHr2 = "--------------------------------------------------";



    def adjustDate(self,date):
        """
        調整日期，如傳入日期為星期六/日，調整為「星期五」.

        Args:
            date (datetime): 要調整的日期
        Returns:
            datetime: 調整後的日期
        """
        # print('\t\t ===== adjustDate() ===== ')
        
        if date.weekday() == 5:
            # 如果傳入的「傳入日期」是「星期六」，調整為「星期五」
            date -= timedelta(days=1)
        elif date.weekday() == 6:
            # 如果傳入的「傳入日期」是「星期日」，調整為「星期五」
            date -= timedelta(days=2)

        # print(F"\t\t date = ({type(date)}) {date}")
        return date



    def getStartDate(self,dEndDate,dStartDate=""):
        """
        取得「開始日期」.

        Args:
            dEndDate (datetime): 結束日期
            sStartDate (str): 開始日期，如傳入空值，會設為「結束日期」的前 31 個交易日.
        Returns:
            datetime: 開始日期.
        """
        # print('\t\t ===== getStartDate() ===== ')
        # print(F"\t\t dEndDate   = {type(dEndDate)} {dEndDate}")
        # print(F"\t\t dStartDate = {type(dStartDate)} {dStartDate}")

        # ------------------------------
        # get 開始日期: 如果沒有傳入「開始日期」，預設為 「結束日期」的前 31 個交易日
        # ------------------------------
        if dStartDate == "":
            # print(F"\t\t get dStartDate {self.sHr1}")

            # 使用 yfinance 獲取台灣加權指數（^TWII）的歷史數據
            # progress=False 表示不要印出進度條，不然執行會印出 [*********************100%%**********************]  1 of 1 completed
            dYahooStart = dEndDate - timedelta(days=90)
            twii = yf.download('^TWII', start=dYahooStart, end=dEndDate, progress=False)
            # print(F"\t\t twii = ({type(twii)}) {len(twii)}")

            # 依日期排序
            twii = twii.sort_index(ascending=False)

            i = 0
            for index, row in twii.iterrows():
                # print(F"\t\t\t {i} index = ({type(index)}) {index}")
                i += 1
                if i >= 31:
                    dStartDate = index.to_pydatetime()
                    break
        # print(F"\t\t dStartDate = ({type(dStartDate)}) {dStartDate}")

        return dStartDate



    def getStockAnalyze(self,stockCode,sEndDate="",sStartDate=""):
        """
        取得「指定股票」的基本資料 及 各項技術指標
         
        (TWSE 有 request limit, 每 5 秒鐘 3 個 request，超過的話會被 ban 掉，請自行注意)

        Args:
            stockCode (str) : 股票代碼
            sEndDate (str)  : 取得交易日期區間的「結束日期」，如傳入空值，預設為「今日」，格式: 2024-06-20
            sStartDate (str): 取得交易日期區間的「開始日期」，如傳入空值，預設為「今日-31」，格式: 2024-06-20
        Returns:
            list
                index[0] : 資料日期.
                index[1] : 股票代號.
                index[2] : 股票名稱.
                index[3] : 收盤價.
                index[4] : 多頭數量.
                index[5] : 空頭數量.
                index[6] : 詳細各計術指標值、歷史價格數據.
        """
        print('\t ===== getStockAnalyze() ===== ')
        print(F"\t stockCode  = ({type(stockCode)}) {stockCode}")
        print(F"\t sStartDate = ({type(sStartDate)}) {sStartDate}")
        print(F"\t sEndDate   = ({type(sEndDate)}) {sEndDate}")
        listResult = []


        # ------------------------------
        # get 股票基本資料 (twstock)
        # ------------------------------
        # 檢查 stockCode 是否是字串，如果不是，將其轉換為字串
        if not isinstance(stockCode, str):
            stockCode = str(stockCode)

        # new twstock.Stock 類型的物件，用於後續取得股票資料 <class 'twstock.stock.Stock'>
        tObjStock = twstock.Stock(stockCode)

        # 取得股票基本資料 ex: (type='股票', code='6197', name='佳必琪', ISIN='TW0006197007', start='2004/11/08', market='上市', group='電子零組件業', CFI='ESVUFR')
        stockInfo = twstock.codes.get(stockCode)
        stockName = stockInfo.name
        print(F"\t stockName = ({type(stockName)}) {stockName}")


        # ------------------------------
        # set 結束日期
        # 1.如果沒有傳入「結束日期」，結束日期 = 當天
        # 2.如果  有傳入「結束日期」，結束日期 = 傳入結束日期
        # 皆調整日期不得為六、日
        # ------------------------------
        if sEndDate != "":
            # string 轉成 datetime.datetime 並 
            dEndDate = datetime.strptime(sEndDate, '%Y-%m-%d')
        else:
            # 如果沒有傳入「結束日期」，預設為 今日
            dEndDate = self.dToday
        # 調整日期不得為六、日
        dEndDate = self.adjustDate(dEndDate)
        print(F"\t dEndDate = ({type(dEndDate)}) {dEndDate}")


        # ------------------------------
        # set 開始日期
        # 1.如果沒有傳入「開始日期」，開始日期 = 結束日期 - 31個交易日
        # 2.如果  有傳入「開始日期」
        #   如果「傳入開始日期」<「結束日期 - 31個交易日」，開始日期 = 結束日期 - 31個交易日
        #   如果「傳入開始日期」>「結束日期 - 31個交易日」，開始日期 = 傳入開始日期，調整日期不得為六、日
        # ------------------------------
        # get 「結束日期」的前 31 個交易日
        dStartDate = self.getStartDate(dEndDate,"")
        
        # 抓取資料的開始日期 = 開始日期 - 31個交易日
        if sStartDate != "":
            # string 轉成 datetime.datetime
            dInStartDate = datetime.strptime(sStartDate, '%Y-%m-%d')
            if dInStartDate < dStartDate:
                dInStartDate = self.adjustDate(dInStartDate)
                dStartDate = dInStartDate
        print(F"\t dStartDate = ({type(dStartDate)}) {dStartDate}")


        # ------------------------------
        # set 股票歷史數據開始日期 = 開始日期 - 31個交易日
        # ------------------------------
        dExtendedStartDate = self.getStartDate(dStartDate,"")
        print(F"\t dExtendedStartDate = ({type(dExtendedStartDate)}) {dExtendedStartDate}")


        # ------------------------------
        # set 股票歷史數據
        # ------------------------------
        # get 交易數據 <class 'numpy.ndarray'>
        # 如果沒有先用 fetch_from 取到歷史資料，預設抓 31 天的資料
        # npPrice  = np.array(tObjStock.price   , dtype='float64') # 歷史價格
        # npHigh   = np.array(tObjStock.high    , dtype='float64') # 歷史最高價
        # npLow    = np.array(tObjStock.low     , dtype='float64') # 歷史最低價
        # npClose  = np.array(tObjStock.close   , dtype='float64') # 歷史收盤價
        # npVolume = np.array(tObjStock.capacity, dtype='float64') # 歷史成交量
        # print(F"\t npPrice = ({type(npPrice)}) (size ={len(npPrice)}) {npPrice}")

        # get 股票歷史數據: 從 dExtendedStartDate 年和月開始的歷史數據
        listHistoryData = tObjStock.fetch_from(dExtendedStartDate.year, dExtendedStartDate.month)
        print(F"\t listHistoryData = ({type(listHistoryData)}) (size ={len(listHistoryData)})")


        # ------------------------------
        # 迴圈讀取「開始日期-結束日期」的每一個交易日，並依不同日期，取對應區間的數據計算技術指標
        # 例如:
        # 2024-07-12 歷史數據區間為 2024-07-12 到 2024-05-29 (2024-07-12 前31天)
        # 2024-07-11 歷史數據區間為 2024-07-11 到 2024-05-28 (2024-07-11 前31天)
        # ------------------------------
        # modify by huey,2024/07/13,太多筆的時候，如果一直從yahoo取日期，會有連線錯誤的問題，故改寫
        # set 股票歷史數據開始日期 = 開始日期
        # 使用 yfinance 獲取台灣加權指數（^TWII）的歷史數據，取得「開始日期-結束日期」的交易日
        # dYahooEndDate = dEndDate - timedelta(days=-1)
        # twii = yf.download('^TWII', start=dStartDate, end=dYahooEndDate, progress=False)
        # print(F"\t twii = ({type(twii)}) (size ={len(twii)})")

        # 驗證 yahoo 取到的交易日與「開始日期-結束日期」符合
        # max_date = twii.index.max() # 最大交易日
        # min_date = twii.index.min() # 最小交易日
        # print(F"\t max_date = ({type(max_date)}) {max_date}")
        # print(F"\t min_date = ({type(min_date)}) {min_date}")


        # ===== 先取到「每個交易日」的前 31 個交易日 =====
        dictHistStartDate = {}

        # <class 'list'> 存放的資料類型為 datetime，將 List 由大至小排序
        listTradingDates = sorted(tObjStock.date, reverse=True)

        iDateNo = 0
        for dCurrentDate in listTradingDates:
            iDateNo += 1
            sCurrentDate = dCurrentDate.strftime('%Y-%m-%d')

            # 如果日期小於開始日期，跳過此次迴圈
            if dCurrentDate < dStartDate:
                continue
            # print(F"\t {iDateNo}.dCurrentDate = {dCurrentDate} {self.sHr1}")

            i = 0
            for index in listTradingDates:
                # ex: index = (<class 'datetime.datetime'>) 2024-07-12 00:00:00

                if dCurrentDate < index:
                    continue
                elif dCurrentDate == index:
                    i = 0
                i += 1
                # print(F"\t\t\t {i} index = ({type(index)}) {index}")

                if i == 32:
                    # print(F"\t\t\t {i} index = ({type(index)}) {index}")
                    dictHistStartDate[dCurrentDate] = index
                    continue


        # ===== 迴圈讀取「每個交易日」 =====
        # 按 key 升序排序
        dictHistStartDate = dict(sorted(dictHistStartDate.items()))

        iDateNo = 0
        # for index, row in twii.iterrows():
        for dCurrentDate, dHistStartDate in dictHistStartDate.items():
            iDateNo += 1
            sCurrentDate = dCurrentDate.strftime('%Y-%m-%d')
            sHistStartDate = dHistStartDate.strftime('%Y-%m-%d')

            # dCurrentDate = index.to_pydatetime()
            # sCurrentDate = dCurrentDate.strftime('%Y-%m-%d')

            # # get 「當前日期」的前 31 個交易日
            # dHistStartDate = self.getStartDate(dCurrentDate,"")
            # sHistStartDate = dHistStartDate.strftime('%Y-%m-%d')

            # print(F"\t {self.sHr1}")
            # print(F"\t {iDateNo}.sCurrentDate = {sCurrentDate}，計算數據區間為 {sCurrentDate} 到 {sHistStartDate}")

            # 篩選出指定日期範圍內的數據，並存入 list 中
            listDate   = [] # 交易日
            listPrice  = [] # 歷史價格
            listHigh   = [] # 歷史最高價
            listLow    = [] # 歷史最低價
            listClose  = [] # 歷史收盤價
            listVolume = [] # 歷史成交量

            for d, p, h, l, c, v in zip(tObjStock.date, tObjStock.price, tObjStock.high, tObjStock.low, tObjStock.close, tObjStock.capacity):
                if dHistStartDate <= d <= dCurrentDate:
                    listDate.append(d)
                    listPrice.append(p)
                    listHigh.append(h)
                    listLow.append(l)
                    listClose.append(c)
                    listVolume.append(v)

            # 將 list 轉換為 NumPy 數組，每個數組長度為 32
            npHigh   = np.array(listHigh  , dtype='float64') # 歷史最高價
            npLow    = np.array(listLow   , dtype='float64') # 歷史最低價
            npClose  = np.array(listClose , dtype='float64') # 歷史收盤價
            npVolume = np.array(listVolume, dtype='float64') # 歷史成交量
            # print(F"\t\t 歷史最高價 npHigh   = ({type(npHigh)}) (size ={len(npHigh)})")
            # print(F"\t\t 歷史最低價 npLow    = ({type(npLow)}) (size ={len(npLow)})")
            # print(F"\t\t 歷史收盤價 npClose  = ({type(npClose)}) (size ={len(npClose)})")
            # print(F"\t\t 歷史成交量 npVolume = ({type(npVolume)}) (size ={len(npVolume)})")



            # ------------------------------
            # 計算各項技術指標
            # ------------------------------
            # 三重指數平滑異同移動平均線 (TRIX)、三重平滑指標 (TEMA): 皆取到nan，導致該項計算皆為「空頭」，故移除該項
            # modify by huey,2024/06/30,註解: 移動平均收斂背離指標 (MACD)
            # 會因為歷史資料數不同而取到不同的值，似乎日期前 33 個都取不到值，如果要用此指標，目前的程式不適用，故暫註解，日後要用再修改
            # modify by huey,2024/06/30,註解: 平均方向性指數 (ADX)
            # 會因為歷史資料數不同而取到不同的值，故暫註解，日後要用再修改
            # modify by huey,2024/07/02,先將「移動平均收斂背離指標 (MACD)、平均方向性指數 (ADX)」加回來
            # modify by huey,2024/07/05,「Price and Volume Trend (PVT)」與「平衡交易量 (OBV)」重覆
            # modify by huey,2024/07/09,「Ease of Movement (EMV)」一直判斷為「多頭」不準，故移除該項
            # modify by huey,2024/07/09,「標準差 (STDDEV)」跟「終極振盪器 (ULTOSC)」與各股漲跌區勢不符合，故移除該項
            sma = talib.SMA(npClose, timeperiod=5) #簡單移動平均線 (SMA)
            ema = talib.EMA(npClose, timeperiod=5) #指數移動平均線 (EMA)
            macd, macdsignal, macdhist = talib.MACD(npClose, fastperiod=12, slowperiod=26, signalperiod=9) #移動平均收斂背離指標 (MACD)
            rsi = talib.RSI(npClose, timeperiod=14) #相對強弱指數 (RSI)
            slowk, slowd = talib.STOCH(npHigh, npLow, npClose, fastk_period=14, slowk_period=3, slowk_matype=0, slowd_period=3, slowd_matype=0) #隨機指標 (STOCH)
            upperband, middleband, lowerband = talib.BBANDS(npClose, timeperiod=20, nbdevup=2, nbdevdn=2, matype=0) #布林帶 (BBANDS)
            obv    = talib.OBV(npClose, npVolume) #平衡交易量 (OBV)
            willr  = talib.WILLR(npHigh, npLow, npClose, timeperiod=14) #威廉指標 (WILLR)
            atr    = talib.ATR(npHigh, npLow, npClose, timeperiod=14) #平均真實範圍 (ATR)
            cci    = talib.CCI(npHigh, npLow, npClose, timeperiod=14) #商品通道指數 (CCI)
            mom    = talib.MOM(npClose, timeperiod=10) #動量指標 (MOM)
            sar    = talib.SAR(npHigh, npLow, acceleration=0.02, maximum=0.2) #龐氏指標 (SAR)
            adx    = talib.ADX(npHigh, npLow, npClose, timeperiod=14) #平均方向性指數 (ADX)
            # trix = talib.TRIX(npClose, timeperiod=30) #三重指數平滑異同移動平均線 (TRIX)
            # tema = talib.TEMA(npClose, timeperiod=30) #三重平滑指標 (TEMA)
            stddev = talib.STDDEV(npClose, timeperiod=5, nbdev=1) #股票或其他金融時間序列資料的標準差 (Standard Deviation)
            varma  = talib.VAR(npClose, timeperiod=14, nbdev=1) #可變移動平均線 (VARMA)
            # print(F"簡單移動平均線 (SMA) = {type(sma)} {sma}")
            # print(F"指數移動平均線 (EMA) = {type(ema)} {ema}")
            # print(F"移動平均收斂背離指標 (MACD) : macd = {type(macd)} {macd}")
            # print(F"移動平均收斂背離指標 (MACD) : macdsignal = {type(macdsignal)} {macdsignal}")
            # print(F"移動平均收斂背離指標 (MACD) : macdhist = {type(macdhist)} {macdhist}")
            # print(F"相對強弱指數 (RSI) = {type(rsi)} {rsi}")
            # print(F"隨機指標 (STOCH) : slowk = {type(slowk)} {slowk},slowd = {type(slowd)} {slowd}")
            # print(F"布林帶 (BBANDS) : upperband = {type(upperband)} {upperband},middleband = {type(middleband)} {middleband},lowerband = {type(lowerband)} {lowerband}")
            # print(F"平衡交易量 (OBV) = {type(obv)} {obv}")
            # print(F"威廉指標 (WILLR) = {type(willr)} {willr}")
            # print(F"平均真實範圍 (ATR) = {type(atr)} {atr}")
            # print(F"商品通道指數 (CCI) = {type(cci)} {cci}")
            # print(F"動量指標 (MOM) = {type(mom)} {mom}")
            # print(F"龐氏指標 (SAR) = {type(sar)} {sar}")
            # print(F"平均方向性指數 (ADX) = {type(adx)} {adx}")
            # print(F"三重指數平滑異同移動平均線 (TRIX) = {type(trix)} {trix}")
            # print(F"三重平滑指標 (TEMA) = {type(tema)} {tema}")
            # print(F"股票或其他金融時間序列資料的標準差 (Standard Deviation) = {type(stddev)} {stddev}")
            # print(F"可變移動平均線 (VARMA) = {type(varma)} {varma}")

            # 計算VMA (Volume Moving Average)
            vma_short = talib.SMA(npVolume, timeperiod=5)
            vma_mid   = talib.SMA(npVolume, timeperiod=10)
            vma_long  = talib.SMA(npVolume, timeperiod=20)
            # print(F"成交量移動平均線 (VMA 短期) vma_short = {type(vma_short)} {vma_short}")
            # print(F"成交量移動平均線 (VMA 中期) vma_mid   = {type(vma_mid)} {vma_mid}")
            # print(F"成交量移動平均線 (VMA 長期) vma_long  = {type(vma_long)} {vma_long}")

            # 計算VROC (Volume Rate of Change)
            vroc = talib.ROC(npVolume, timeperiod=12)
            # print(F"成交量比率 (VROC) vroc = {type(vroc)} {vroc}")

            # 力量指標 (FORCE) 需要自訂計算
            def calc_force_index(close, volume, period=13):
                force = talib.ROC(close, 1) * volume
                return talib.SMA(force, timeperiod=period)
            force_index = calc_force_index(npClose, npVolume, period=13)
            # print(F"力量指標 (FORCE) = {type(force_index)} {force_index}")

            # 標準差 (STDDEV)
            # upper_stddev, lower_stddev = middleband + stddev, middleband - stddev
            # print(F"標準差 (STDDEV) : upper_stddev = {type(upper_stddev)} {upper_stddev},lower_stddev = {type(lower_stddev)} {lower_stddev}")

            # 其他指標計算
            cmo  = talib.CMO(npClose, timeperiod=14) #Chande Momentum Oscillator (CMO)
            # emv  = talib.LINEARREG(npHigh - npLow, timeperiod=14) #Ease of Movement (EMV)
            kama = talib.KAMA(npClose, timeperiod=30) #Kaufman Adaptive Moving Average (KAMA)
            mfi  = talib.MFI(npHigh, npLow, npClose, npVolume, timeperiod=14) #Money Flow Index (MFI)
            # print(F"cmo  = {type(cmo)} {cmo}")
            # print(F"emv  = {type(cmo)} {emv}")
            # print(F"kama = {type(cmo)} {kama}")
            # print(F"mfi  = {type(mfi)} {mfi}")

           # add by huey,2024/07/07,加入新的技術指標
            linearreg       = talib.LINEARREG(npClose, timeperiod=14)       # 線性回歸 (LINEARREG)
            linearreg_angle = talib.LINEARREG_ANGLE(npClose, timeperiod=14) #線性回歸角度 (LINEARREG_ANGLE)
            linearreg_slope = talib.LINEARREG_SLOPE(npClose, timeperiod=14) #線性回歸斜率 (LINEARREG_SLOPE)
            medprice = talib.MEDPRICE(npHigh, npLow) #中價 (MEDPRICE)
            ppo  = talib.PPO(npClose, fastperiod=12, slowperiod=26, matype=0) #Percentage Price Oscillator (PPO)
            roc  = talib.ROC(npClose, timeperiod=10)  #Rate of change (ROC)
            rocp = talib.ROCP(npClose, timeperiod=10) #Rate of change Percentage (ROCP)
            tsf  = talib.TSF(npClose, timeperiod=14)  #時間序列預測 (TSF)
            typprice = talib.TYPPRICE(npHigh, npLow, npClose)  #典型價格 (TYPPRICE)
            # ultosc = talib.ULTOSC(npHigh, npLow, npClose, timeperiod1=7, timeperiod2=14, timeperiod3=28)  #終極振盪器 (ULTOSC)
            wclprice = talib.WCLPRICE(npHigh, npLow, npClose)  #加權收盤價格 (WCLPRICE)

            for i in range(len(listDate)):
                dDate   = listDate[i]   # 交易日 <class 'datetime.datetime'>
                fPrice  = listPrice[i]  # 歷史價格   <class 'float'>
                fHigh   = listHigh[i]   # 歷史最高價 <class 'float'>
                fLow    = listLow[i]    # 歷史最低價 <class 'float'>
                fClose  = listClose[i]  # 歷史收盤價 <class 'float'>
                fVolume = listVolume[i] # 歷史成交量 <class 'int'>

                sDate = dDate.strftime('%Y-%m-%d')

                # 這次讀到的日期 才計算技術指標為「多頭」還是「空頭」
                if sCurrentDate != sDate:
                    continue
                # print(F"\t\t {i}.{sDate} {self.sHr2}")

                nClosePrice = npClose[i]   # 歷史收盤價 <class 'numpy.float64'>
                nVolume     = npVolume[i]  # 歷史成交量 <class 'numpy.float64'>
                # print(F"\t\t nClosePrice = {type(nClosePrice)} {nClosePrice}")
                # print(F"\t\t nVolume = {type(nVolume)} {nVolume}")

                iUpCount   = 0   #多頭數量
                iDownCount = 0   #空頭數量
                dictSignals = {} #記錄每個技術指標 多頭 vs 空頭
                # 宣告dictionary，記錄每個技術指標 多頭 vs 空頭
                dictSignals = {
                    "簡單移動平均線 (SMA)"        : "多頭" if nClosePrice > sma[i] else "空頭",
                    "指數移動平均線 (EMA)"        : "多頭" if nClosePrice > ema[i] else "空頭",
                    "移動平均收斂背離指標 (MACD)" : "多頭" if macd[i] > macdsignal[i] else "空頭",
                    "相對強弱指數 (RSI)"          : "多頭" if rsi[i] > 50 else "空頭",
                    "平衡交易量 (OBV)"            : "多頭" if obv[i] > obv[i-1] else "空頭",
                    "威廉指標 (WILLR)"            : "多頭" if willr[i] > -50 else "空頭",
                    "動量指標 (MOM)"              : "多頭" if mom[i] > 0 else "空頭",
                    "龐氏指標 (SAR)"              : "多頭" if nClosePrice > sar[i] else "空頭",
                    "力量指標 (FORCE)"            : "多頭" if force_index[i] > 0 else "空頭",
                    "成交量比率 (VROC)"           : "多頭" if vroc[i] > 0 else "空頭",
                    "Money Flow Index (MFI)"      : "多頭" if mfi[i] > 50 else "空頭",
                    # add by huey,2024/07/07,加入新的技術指標
                    "線性回歸 (LINEARREG)"          : "多頭" if linearreg[i] > linearreg[i-1] else "空頭",
                    "線性回歸角度 (LINEARREG_ANGLE)": "多頭" if linearreg_angle[i] > linearreg_angle[i-1] else "空頭",
                    "線性回歸斜率 (LINEARREG_SLOPE)": "多頭" if linearreg_slope[i] > linearreg_slope[i-1] else "空頭",
                    "中價 (MEDPRICE)"               : "多頭" if medprice[i] > medprice[i-1] else "空頭",
                    "Percentage Price Oscillator (PPO)": "多頭" if ppo[i] > 0 else "空頭",
                    "Rate of change (ROC)"             : "多頭" if roc[i] > 0 else "空頭",
                    "Rate of change Percentage (ROCP)" : "多頭" if rocp[i] > 0 else "空頭",
                    "時間序列預測 (TSF)"               : "多頭" if tsf[i] > tsf[i-1] else "空頭",
                    "典型價格 (TYPPRICE)"              : "多頭" if typprice[i] > typprice[i-1] else "空頭",
                    # "終極振盪器 (ULTOSC)"            : "多頭" if ultosc[i] > 50 else "空頭",
                    "加權收盤價格 (WCLPRICE)"          : "多頭" if wclprice[i] > wclprice[i-1] else "空頭",
                    # add by huey,2024/07/07,移除的技術指標
                    "隨機指標 (STOCH)"                      : "多頭" if slowk[i] > slowd[i] else "空頭",
                    "布林帶 (BBANDS)"                       : "多頭" if nClosePrice < upperband[i] and nClosePrice > lowerband[i] else "空頭",
                    "平均真實範圍 (ATR)"                    : "多頭" if atr[i] > atr[i-1] else "空頭",
                    "商品通道指數 (CCI)"                    : "多頭" if cci[i] > 100 else "空頭",
                    # "標準差 (STDDEV)"                       : "多頭" if nClosePrice < upper_stddev[i] and nClosePrice > lower_stddev[i] else "空頭",
                        "平均方向性指數 (ADX)"                 : "多頭" if adx[i] > 25 else "空頭",
                    "可變移動平均線 (VARMA)"                : "多頭" if varma[i] > varma[i-1] else "空頭",
                    "成交量移動平均線 (VMA 短期)"           : "多頭" if nVolume > vma_short[i] else "空頭",
                    "成交量移動平均線 (VMA 中期)"           : "多頭" if nVolume > vma_mid[i] else "空頭",
                    "成交量移動平均線 (VMA 長期)"           : "多頭" if nVolume > vma_long[i] else "空頭",
                    "Chande Momentum Oscillator (CMO)"      : "多頭" if cmo[i] > 0 else "空頭",
                    # "Ease of Movement (EMV)"              : "多頭" if emv[i] > 0 else "空頭",
                    "Kaufman Adaptive Moving Average (KAMA)": "多頭" if nClosePrice > kama[i] else "空頭",
                }
                # print(F"\t\t dictSignals = ({type(dictSignals)}) (size ={len(dictSignals)})")

                # 計算「多頭」和「空頭」的數量
                iUpCount   = sum(1 for signal in dictSignals.values() if signal == "多頭")
                iDownCount = sum(1 for signal in dictSignals.values() if signal == "空頭")
                # print(F"\t\t iUpCount   = ({type(iUpCount)}) {iUpCount}")
                # print(F"\t\t iDownCount = ({type(iDownCount)}) {iDownCount}")

                # 將 「多頭」和「空頭」的數量 加入到 dictSignals
                dictSignals["多頭數量"] = iUpCount
                dictSignals["空頭數量"] = iDownCount

                # 基本訊息也加入
                dictSignals["歷史價格"]   = fPrice  # 歷史價格   <class 'float'>
                dictSignals["歷史最高價"] = fHigh   # 歷史最高價 <class 'float'>
                dictSignals["歷史最低價"] = fLow    # 歷史最低價 <class 'float'>
                dictSignals["歷史收盤價"] = fClose  # 歷史收盤價 <class 'float'>
                dictSignals["歷史成交量"] = fVolume # 歷史成交量 <class 'int'>

                # Set 要回傳的list
                # index[0] # 資料日期
                # index[1] # 股票代號
                # index[2] # 股票名稱
                # index[3] # 收盤價
                # index[4] # 多頭數量
                # index[5] # 空頭數量
                # index[6] # 詳細各計術指標值、歷史價格數據
                listResult.append([sDate,stockCode,stockName,fClose,iUpCount,iDownCount,dictSignals])
                # print(F"\t\t listResult = ({type(listResult)}) (size ={len(listResult)})")


        # 將 List 用日期降序排序
        listResult.sort(key=lambda x: x[0], reverse=True)

        # get 「多頭/空頭數量」連續增加天數
        dictDays = self.countIncreaseDaysHistory(listResult)

        return listResult


    def countIncreaseDaysHistory(self,listStockInfo):
        '''
        針對傳入的list，每一天都取得「多頭/空頭數量」連續增加天數，回傳 dict
        
        例如:
        2024-06-26 多頭數量: 16,
        2024-06-25 多頭數量: 15,
        2024-06-24 多頭數量: 15,
        2024-06-21 多頭數量: 13,
        2024-06-20 多頭數量: 11,
       26日的「多頭 連續增加天數 (2024-06-26Up)」, 將回傳 4

        Args:
            listStockInfo (list): 股票資訊List
        Returns:
            dict:
                KEY: stockDate+"Up"   = 指定股票日期的「多頭 連續增加天數」
                KEY: stockDate+"Down" = 指定股票日期的「空頭 連續增加天數」
        '''
        # print('\t ===== countIncreaseDays() ===== ')
        ilistStockInfo = len(listStockInfo)
        # print(F"\t listStockInfo size = {type(listStockInfo)} {ilistStockInfo}")

        # Set 要回傳的天數，預為 0
        dictDays = {}

        for a in range(ilistStockInfo):
            # 如果已經比到 list 的最後一筆，結束迴圈
            # if i == ilistStockInfo-1:
            #     break

            # ------------------------------
            # get 「此次」迴圈的日期資料 (ex:2024-06-26)
            # ------------------------------
            listStock = listStockInfo[a]
            stockDate     = listStock[0] # 資料日期
            stockCode     = listStock[1] # 股票代號
            stockName     = listStock[2] # 股票名稱
            fClosePrice   = listStock[3] # 收盤價
            # print(F"\t [countIncreaseDays] a = {a} {self.sHr1}")
            # print(f"\t stockDate = {stockDate},listStock = {listStock}")

            dstockDate = datetime.strptime(stockDate, '%Y-%m-%d')

            iUpDays = 0
            iDownDays = 0
            endUp = True
            endDown = True

            for i in range(ilistStockInfo):
                # 如果已經比到 list 的最後一筆，結束迴圈
                if i == ilistStockInfo-1:
                    break

                # <class 'list'>
                listCurrent = listStockInfo[i]
                listPreviou = listStockInfo[i+1]

                # ------------------------------
                # get 「此次」迴圈的日期資料 (ex:2024-06-26)
                # ------------------------------
                sCurrentDate    = listCurrent[0] # 資料日期
                # stockCode     = listCurrent[1] # 股票代號
                # stockName     = listCurrent[2] # 股票名稱
                # fClosePrice   = listCurrent[3] # 收盤價
                iCurrentUpQty   = listCurrent[4] # 多頭數量
                iCurrentDownQty = listCurrent[5] # 空頭數量
                dateCurrent = datetime.strptime(sCurrentDate, '%Y-%m-%d')

                if dateCurrent > dstockDate:
                    continue
                # print(F"\t\t i = {i} {self.sHr1}")
                # print(f"\t\t listCurrent = {listCurrent}, 上一筆 = {listPreviou}")

                # ------------------------------
                # get 「前一天」迴圈的日期資料 (ex:2024-06-25)
                # ------------------------------
                iPreviouUpQty   = listPreviou[4] # 多頭數量
                iPreviouDownQty = listPreviou[5] # 空頭數量
                # dictSignals   = listCurrent[6] # 詳細各計術指標值、歷史價格數據

                # ------------------------------
                # 比較 「此次日期 的 多頭/空頭數數量」 >= 「前一天日期 的 多頭/空頭數數量」，連續天數加一
                # ------------------------------
                if endUp and iCurrentUpQty >= iPreviouUpQty:
                    iUpDays += 1
                else:
                    endUp = False
                #     break
                # print(f"\t\t 多頭數量連續:  Current = {iCurrentUpQty}, Previou = {iPreviouUpQty}, iUpDays = {iUpDays} ({endUp})")

                if endDown and iCurrentDownQty >= iPreviouDownQty:
                    iDownDays += 1
                else:
                    endDown = False
                #     break
                # print(f"\t\t 空頭數量連續:  Current = {iCurrentDownQty}, Previou = {iPreviouDownQty}, iDownDays = {iDownDays} ({endDown})")

                if endUp == False and endDown == False:
                    break

            dictDays[stockDate+"Up"] = iUpDays
            dictDays[stockDate+"Down"] = iDownDays
            # print(f"\t iUpDays = {iUpDays},iDownDays = {iDownDays}")

        return dictDays


    def countIncreaseDays(self,sType,listStockInfo):
        '''
        只取得「最新日期」的「多頭/空頭數量」連續增加天數，回傳 int
        
        例如: 06/19,06/20,06/21 的「多頭/空頭空數量」分別為 15,15,16, 將回傳 2

        Args:
            sType (str): 請傳入 up / down, up: 代表要計算「多頭」; down: 代表要計算「空頭」
            listStockInfo (list): 股票資訊List
        Returns:
            int
        '''
        # print('\t ===== countIncreaseDays() ===== ')
        # print(F"\t sType = {type(sType)} {sType}")
        # print(F"\t listStockInfo size = {type(listStockInfo)} {len(listStockInfo)}")

        # 如果傳入的 list 為空，直接回傳 0
        if not listStockInfo:
            return 0

        # Set 要回傳的天數，預為 0
        iCountDays = 0

        ilistStockInfo = len(listStockInfo)

        # 遍歷 stockinfo 從第二天開始到最後一天
        for i in range(len(listStockInfo)):
            # print(F"\t i = {i} {self.sHr1}")

            # 如果已經比到 list 的最後一筆，結束迴圈
            if i == ilistStockInfo-1:
                break

            # <class 'list'>
            listCurrent = listStockInfo[i]
            listPreviou = listStockInfo[i+1]

            # !!方便測試時，提早跳出迴圈
            # if i == 3:
            #     break
            # print(F"\t listCurrent = {listCurrent} ,listPreviou = {listPreviou}")

            # get 這次迴圈日期 & 前一天的日期的「多頭數量」
            sDate = listCurrent[0]
            if sType == "down":
                iCurrentQty  = listCurrent[5]
                iPreviousQty = listPreviou[5]
            else:
                iCurrentQty  = listCurrent[4]
                iPreviousQty = listPreviou[4]
            # print(F"\t i = {i},sDate = {sDate},iCurrentQty = {iCurrentQty}, iPreviousQty = {iPreviousQty}")

            # 如果 這次迴圈日期的多頭數是 > 前一天的，連續天數加一
            if iCurrentQty > iPreviousQty:
                iCountDays += 1
                # print(F"\t iCountDays = {iCountDays}")
            else:
                break

        return iCountDays

    def countIndexDays(self,sType,listStockInfo,index=20):
        '''
        取得「多頭/空頭數量」連續 >=index 的天數

        例如: 06/19,06/20,06/21 的「多頭/空頭空數量」分別為 10,15,16, index = 15, 將回傳 2

        Args:
            sType (str): 請傳入 up / down, up: 代表要計算「多頭」; down: 代表要計算「空頭」
            listStockInfo (list): 股票資訊List
            index (int): 「多頭/空頭數量」標準，如傳入空值，預設為 20
        '''
        # print('\t ===== countIndexDays() ===== ')
        # print(F"\t listStockInfo size = {type(listStockInfo)} {len(listStockInfo)}")
        # print(F"\t index = {type(index)} {index}")

        # 如果傳入的 list 為空，直接回傳 0
        if not listStockInfo:
            return 0

        # Set 要回傳的天數，預為 0
        iCountDays = 0

        # 遍歷 stockinfo 從第二天開始到最後一天
        for i in range(len(listStockInfo)):
            # <class 'list'>
            listCurrentDay  = listStockInfo[i]

            # !!方便測試時，提早跳出迴圈
            # if i == 3:
            #     break
            # print(F"\t i = ({i} {self.sHr1}")
            # print(F"\t listCurrentDay  = {i}.{listCurrentDay}")

            # get 這次迴圈日期 & 前一天的日期的「多頭數量」
            sDate = listCurrentDay [0]
            if sType == "down":
                iCurrentQty  = listCurrentDay[5]
            else:
                iCurrentQty  = listCurrentDay[4]
            # print(F"\t i = {i},sDate = {sDate},iCurrentQty = {iCurrentQty}")

            # 如果 這次迴圈日期的多頭數是 > index，連續天數加一
            if iCurrentQty > index:
                iCountDays += 1
                # print(F"\t iCountDays = {iCountDays}")
            else:
                break
        return iCountDays


    def saveToExcel(self, sPath, listStockInfo, dictDays):
        '''
        將數據保存到 Excel 文件中

        Args:
            sPath (str): 文件保存的路徑
            listStockInfo (list): 股票資訊List
            dictDays (dict): 「多頭/空頭數量」連續增加天數
        '''
        # print('\t ===== saveToExcel() ===== ')
        # print(F"\t sPath = {type(sPath)} {sPath}")
        # print(F"\t listStockInfo size = {type(listStockInfo)} {len(listStockInfo)}")
        # print(F"\t dictDays size = {type(dictDays)} {len(dictDays)}")


        # ------------------------------
        # 組出要寫到 Excel 的內容
        # ------------------------------
        listOutput = []
        sEndDate, sStartDate = "", ""
        all_signals_keys = set()

        listSignalsKey = ["簡單移動平均線 (SMA)","指數移動平均線 (EMA)","移動平均收斂背離指標 (MACD)","相對強弱指數 (RSI)","平衡交易量 (OBV)","威廉指標 (WILLR)","動量指標 (MOM)","龐氏指標 (SAR)","力量指標 (FORCE)","成交量比率 (VROC)","Money Flow Index (MFI)","線性回歸 (LINEARREG)","線性回歸角度 (LINEARREG_ANGLE)","線性回歸斜率 (LINEARREG_SLOPE)","中價 (MEDPRICE)","Percentage Price Oscillator (PPO)","Rate of change (ROC)","Rate of change Percentage (ROCP)","時間序列預測 (TSF)","典型價格 (TYPPRICE)","加權收盤價格 (WCLPRICE)",
                        "隨機指標 (STOCH)","布林帶 (BBANDS)","平均真實範圍 (ATR)","商品通道指數 (CCI)","平均方向性指數 (ADX)","可變移動平均線 (VARMA)","成交量移動平均線 (VMA 短期)","成交量移動平均線 (VMA 中期)","成交量移動平均線 (VMA 長期)","Chande Momentum Oscillator (CMO)","Kaufman Adaptive Moving Average (KAMA)"
                        ]

        listStockInfo.sort(key=lambda x: x[0], reverse=False)
        for i in range(len(listStockInfo)):
            # <class 'list'>
            listCurrent = listStockInfo[i]
            stockDate   = listCurrent[0] # 資料日期
            stockCode   = listCurrent[1] # 股票代號
            stockName   = listCurrent[2] # 股票名稱
            fClosePrice = listCurrent[3] # 收盤價
            iUpQty      = listCurrent[4] # 多頭數量
            iDownQty    = listCurrent[5] # 空頭數量
            dictSignals = listCurrent[6] # 詳細各計術指標值、歷史價格數據
            iUpDays00   = dictDays[stockDate+"Up"]   # 「多頭數量」連續增加天數
            iDownDays00 = dictDays[stockDate+"Down"] # 「空頭數量」連續增加天數

            # print(F"\t [saveToExcel] i = {i} {self.sHr1}")
            # print(f"\t stockDate = {stockDate}")
            # print(F"\t dictSignals = {type(dictSignals)} {len(dictSignals)}")

            # modify by huey,2024/07/09,增加匯出各個技術指標
            # listOutput.append([stockDate,stockCode,stockName,fClosePrice,iUpQty,iDownQty,iUpDays00,iDownDays00])
            row = [self.sToday,stockDate, stockCode, stockName, fClosePrice, iUpQty, iDownQty, iUpDays00, iDownDays00]

            for key in listSignalsKey:
                value = dictSignals.get(key, "Key不存在")
                row.append(value);
                # print(F"\t [saveToExcel] {key} = {value}")

            listOutput.append(row)

            # 記錄第一筆的日期
            if i == 0:
                sStartDate = stockDate

            # 記錄最後一筆的日期
            if i == len(listStockInfo) - 1:
                sEndDate = stockDate
        # print(F"\t listOutput size = {type(listOutput)} {len(listOutput)}")


        # ------------------------------
        # 將資料寫入 Excel
        # ------------------------------
        # Set 表頭
        columns = ["匯出日期","交易日期", "股票代號", "股票名稱", "收盤價", "技術指標-多頭數量", "技術指標-空頭數量","「多頭數量」連續增加天數","「空頭數量」連續增加天數"] + listSignalsKey
        # print(F"columns type = {type(columns)}")

        df = pd.DataFrame(listOutput, columns=columns)

        # Set 檔案名稱
        sFileName = f"{stockCode}_{stockName}_{sStartDate}_to_{sEndDate}_FromClass.xlsx"

        # 構建完整的文件路徑
        sFilePath = f"{sPath}/{sFileName}"

        # 將資料保存到 Excel 文件
        df.to_excel(sFilePath, index=False)
        print(f"\t 1.資料已保存到 = {sFilePath}")

        # ------------------------------
        # 使用 openpyxl 調整格式
        # ------------------------------
        workbook = load_workbook(sFilePath)
        worksheet = workbook.active

        # 凍結頂端列
        worksheet.freeze_panes = "A2"

        # 設定字體為微軟正黑體
        # font = Font(name='Microsoft JhengHei', size=10)
        # for row in worksheet.iter_rows():
        #     for cell in row:
        #         cell.font = font

        # 設定「第01列」 (worksheet[1] 指的是第一列)
        font = Font(name='Microsoft JhengHei', size=10, bold=True) # 設定 字體 = 微軟正黑體
        fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid') # 設定 單元格背景顏色
        for cell in worksheet[1]:
            cell.alignment = Alignment(wrap_text=True)  # 自動換行
            cell.fill = fill  # 設定背景顏色
            cell.font = font  # 設定字體為微軟正黑體

        # 設定欄寬
        worksheet.column_dimensions['A'].width = 13
        worksheet.column_dimensions['B'].width = 13
        columnsSize10 = ['C', 'D', 'E', 'F', 'G', 'H', 'I']
        for col in columnsSize10:
            worksheet.column_dimensions[col].width = 10

        # 「技術指標」如果是「多頭」，設定背景顏色
        fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        # # 假設 I 列到 AN 列對應的列索引是 9 到 40 (openpyxl 使用0基索引)
        # start_col = 9
        # end_col = 40

        # # 遍歷 I 到 AN 列的每一個單元格
        # for row in worksheet.iter_rows(min_col=start_col, max_col=end_col, min_row=2):
        #     for cell in row:
        #         if cell.value == "多頭":
        #             cell.fill = fill

        # 設定邊框
        thin = Side(border_style="thin", color="000000")
        # border = Border(left=thin, right=thin, top=thin, bottom=thin)
        border = Border(left=thin, right=thin)

        # 設定字體為微軟正黑體
        font = Font(name='Microsoft JhengHei', size=10)

        # 遍歷所有的欄位
        for col_idx, col in enumerate(worksheet.iter_cols(), 1):
            # get 欄號 ex: A,B,C... (<class 'str'>)
            colNo = col[0].column_letter
            # print(F"colNo = ({type(colNo)}) {colNo}")

            # 遍歷每一個欄位中的每一個單元格
            for cell in col[1:]:  # 跳過標題行
                cell.border = border # 設定邊框
                cell.font = font # 設定字體為微軟正黑體
                if cell.value == "多頭":
                    cell.fill = fill

        # 保存更改
        workbook.save(sFilePath)
        print(f"\t 2.資料格式已修改")

    def saveToGoogle(self,listStockInfo, dictDays):
        '''
        將數據保存到 Google Sheet 文件中

        Args:
            listStockInfo (list): 股票資訊List
            path (str): 文件保存的路徑
        '''
        # print('\t ===== saveToGoogle() ===== ')
        # print(F"\t listStockInfo size = {type(listStockInfo)} {len(listStockInfo)}")


        # ------------------------------
        # 前置作業，使用憑證文件，開啟 GoogleSheet
        # ------------------------------
        # get 目前程式檔案路徑
        script_dir = os.path.dirname(__file__)

        # Set JSON 憑證文件的路徑
        credentials_file = os.path.join(script_dir, "stock-426606-5b695171527f.json")

        # 使用 JSON 憑證文件來授權
        gc = pygsheets.authorize(service_file=credentials_file)

        # 開啟 GoogleSheet 試算表(spreadsheet)
        sht  = gc.open_by_key('13ya5frtagmkCojhthWbG82L0NzTshX5PqwK25MNER3k') # 技術指標明細

        # get GoogleSheet 中的第一個 工作表(worksheet)
        wks = sht[0]


        # ------------------------------
        # 組出要寫到 GoogleSheet 的內容
        # ------------------------------
        listOutput = []
        sEndDate, sStartDate = "", ""

        for i in range(len(listStockInfo)):
            listCurrentDay = listStockInfo [i]

            sDate         = listCurrentDay[0]  # 資料日期
            stockCode     = listCurrentDay[1] # 股票代號
            stockName     = listCurrentDay[2] # 股票名稱
            fClosePrice   = listCurrentDay[3] # 收盤價
            iBullishCount = listCurrentDay[4] # 多頭數量
            iBearishCount = listCurrentDay[5] # 空頭數量
            # dictSignals   = listCurrentDay[6] # 詳細各計術指標值、歷史價格數據
            # print(F"\t i = ({i} {self.sHr1}")
            # print(F"\t sDate = {type(sDate)} {sDate}")
            # print(F"\t nClosePrice = {type(nClosePrice)} {nClosePrice}")

            listOutput.append([i,sDate,stockCode,stockName,fClosePrice,iBullishCount,iBearishCount])

            # 記錄第一筆的日期
            if i == 0:
                sStartDate = sDate
            
            # 記錄最後一筆的日期
            if i == len(listStockInfo) - 1:
                sEndDate = sDate
        # print(F"\t listOutput size = {type(listOutput)} {len(listOutput)}")


        # ------------------------------
        # 將資料寫入 GoogleSheet 的內容
        # ------------------------------
        # Set 表頭
        columns = ["i","交易日期", "股票代號", "股票名稱", "收盤價", "技術指標多頭", "技術指標空頭"]
        # print(F"\t columns type = {type(columns)}")

        dfOutput = pd.DataFrame(listOutput, columns=columns)

        # get 目前工作表資料，並轉為 DataFrame
        existingData = wks.get_all_records()
        dfExisting = pd.DataFrame(existingData)

        # 將新的資料附加到現有資料後面
        dfCombined = pd.concat([dfExisting, dfOutput], ignore_index=True).dropna()
        # dfCombined = pd.concat([dfExisting, dfOutput], ignore_index=True)

        # 將更新後的 DataFrame 寫回工作表
        wks.set_dataframe(dfCombined, (0, 0))
        print(f"\t 資料已保存到 = Google Sheet 技術指標明細")


    def getStockData4Yahoo(self,stockCode):
        '''
        爬Yahoo網頁,取得當天股票資訊

        Args:
            stockCode (str) : 股票代碼
        Returns:
            dict
                Key:目前股價.
                Key:漲跌.
                Key:幅度.
                Key:成交量.
                Key:連漲連跌.
        '''
        # print('\t ===== getStockData4Yahoo() ===== ')
        # print(F"\t stockCode = {type(stockCode)} {stockCode}")

        # ------------------------------
        # get 股票基本資料 (Yahoo)
        # ------------------------------
        # 用 yfinance 取得股票資訊，如公司名稱、市值、行業、股息、P/E 比率等
        # yObjStock = yf.Ticker(f"{stockCode}.TW")
        # yStockInfo = yObjStock.info
        # print(F"\t yStockInfo = ({type(yStockInfo)}) {yStockInfo}")

        # get 股票名稱
        # remark by huey,2024/06/22,用 Yahoo 取會取到英文，故註解
        # stockName = yObjStock.info.get('longName', '名稱不可用')
        # print(F"\t stockName = ({type(stockName)}) {stockName}")


        url = 'https://tw.stock.yahoo.com/quote/'+stockCode # Yahoo 股市網址
        web = requests.get(url)                          # 取得網頁內容
        soup = BeautifulSoup(web.text, "html.parser")    # 轉換內容

        divMain = soup.find(id='main-0-QuoteHeader-Proxy')
        # print(F"\t divMain = {divMain}")

        # get 目前股價、漲跌、幅度 DIV
        divPriceElement = divMain.find('div', class_='D(f) Ai(fe) Mb(4px)')
        spanPriceElement = divPriceElement.find_all('span')
        ispanPriceElement = len(spanPriceElement)
        # print(F"\t spanPriceElement  = (Size={ispanPriceElement}) {spanPriceElement}")

        # get 目前股價
        # ex: <span class="Fz(32px) Fw(b) Lh(1) Mend(16px) D(f) Ai(c) C($c-trend-up)">57.3</span>
        spanElement = spanPriceElement[0]
        if spanElement:
            currentPrice = spanElement.text.strip()
        else:
            currentPrice = "取不到，請確認"
        # print(F"\t currentPrice = {currentPrice}")

        # get 漲幅
        spanElement = spanPriceElement[1]
        if spanElement:
            # 檢查是否有 class 屬性，並且 class 屬性中是否包含 'C($c-trend-down)'
            if spanElement.has_attr('class') and 'C($c-trend-down)' in spanElement.get('class'):
                upDown  = "-"+spanPriceElement[1].text
                percent = spanPriceElement[3].text.strip('()')
            elif spanElement.has_attr('class') and 'C($c-trend-up)' in spanElement.get('class'):
                upDown  = "+"+spanPriceElement[1].text
                percent = spanPriceElement[3].text.strip('()')
            else:
                upDown  = spanPriceElement[1].text
                percent = spanPriceElement[2].text.strip('()')
        else:
            upDown = "取不到，請確認"
            percent = "取不到，請確認"
        # print(F"\t upDown = {upDown}")
        # print(F"\t percent = {percent}")

        # get 成交量
        divElement = divMain.find('div', class_='D(f) Fld(c) Ai(c) Fw(b) Pend(8px) Bdendc($bd-primary-divider) Bdends(s) Bdendw(1px)')
        spanElements = divElement.find_all('span')
        if len(spanElements) >= 2:
            qty = spanElements[0].text
        else:
            qty = "取不到，請確認"
        # print(F"\t qty = {qty}")


        # get 連漲連跌
        divElement = divMain.find('div', class_='D(f) Fld(c) Ai(c) Fw(b) Pstart(8px)')
        spanElements = divElement.find_all('span')
        if len(spanElements) >= 2:
            continuous = spanElements[0].text
        else:
            continuous = "取不到，請確認"

        dict = {
            "目前股價": currentPrice,
            "漲跌": upDown,
            "幅度": percent,
            "成交量": qty,
            "連漲連跌": continuous,
        }
        return dict



    # def __getStockData4TWStock(self,stockCode):
        # '''
        # 使用TWStock,取得股票資訊.
        # 私有方法，僅在驗算技術指標使用時才使用，請不要呼叫此Method

        # Args:
            # stockCode (str) : 股票代碼
        # Returns:
            # noen
        # '''
        # print('\t ===== getStockData4TWStock() ===== ')
        # print(F"\t stockCode = {type(stockCode)} {stockCode}")


        # # ------------------------------
        # # get 股票基本資料 (twstock)
        # # ------------------------------
        # # 用 twstock 取得股票資料
        # tObjStock = twstock.Stock(stockCode)

        # # 取得股票基本資料 (type='股票', code='6197', name='佳必琪', ISIN='TW0006197007', start='2004/11/08', market='上市', group='電子零組件業', CFI='ESVUFR')
        # stockInfo = twstock.codes.get(stockCode)
        # stockName = stockInfo.name
        # print(F"\t stockName = ({type(stockName)}) {stockName}")

        # # get 交易日數據 <class 'list'>
        # listDates = tObjStock.date
        # print(F"\t listDates= {type(listDates)} {listDates}")

        # # get 交易數據 <class 'numpy.ndarray'>
        # # 如果沒有先用 fetch_from 取到歷史資料，預設抓 最新交易日至前31 天的資料
        # # EX: 2024/6/28 執行程式，會抓 2024/6/28-2024/5/16
        # npPrice  = np.array(tObjStock.price   , dtype='float64') # 歷史價格
        # npHigh   = np.array(tObjStock.high    , dtype='float64') # 歷史最高價
        # npLow    = np.array(tObjStock.low     , dtype='float64') # 歷史最低價
        # npClose  = np.array(tObjStock.close   , dtype='float64') # 歷史收盤價
        # npVolume = np.array(tObjStock.capacity, dtype='float64') # 歷史成交量
        # print(F"\t 歷史價格   npPrice  = ({type(npPrice)}) {npPrice}")
        # print(F"\t 歷史最高價 npHigh   = ({type(npHigh)}) {npHigh}")
        # print(F"\t 歷史最低價 npLow    = ({type(npLow)}) {npLow}")
        # print(F"\t 歷史收盤價 npClose  = ({type(npClose)}) {npClose}")
        # print(F"\t 歷史成交量 npVolume = ({type(npVolume)}) {npVolume}")

        # # 將日期和價格對應
        # # zip 函數會將 stock_dates 和 stock_prices 這兩個可迭代物件（iterables）打包成一個可迭代的 tuple
        # # 字典生成式的基本語法是 {key: value for item in iterable}
        # # dictPrice  = {date: price for date, price in zip(listDates, npPrice)}
        # # dictHigh   = {date: price for date, price in zip(listDates, npHigh)}
        # # dictLow    = {date: price for date, price in zip(listDates, npLow)}
        # # dictClose  = {date: price for date, price in zip(listDates, npClose)}
        # # dictVolume = {date: price for date, price in zip(listDates, npVolume)}

        # # 印出日期和對應的價格
        # # for date, price in dictPrice.items():
        # #     print(f"Date: {date}, Price: {price}")

        # # ------------------------------
        # # 計算各項技術指標
        # # ------------------------------
        # # 設定技術指標參數
        # iShortTerm = 5 #短期移動平均線的時間周期

        # # 三重指數平滑異同移動平均線 (TRIX)、#三重平滑指標 (TEMA): 皆取到nan，導致該項計算皆為「空頭」，故移除該項
        # sma = talib.SMA(npClose, timeperiod=iShortTerm) #簡單移動平均線 (SMA) <class 'numpy.ndarray'>
        # ema = talib.EMA(npClose, timeperiod=iShortTerm) #指數移動平均線 (EMA) <class 'numpy.ndarray'>
        # macd, macdsignal, macdhist = talib.MACD(npClose, fastperiod=12, slowperiod=26, signalperiod=9) #移動平均收斂背離指標 (MACD)
        # rsi = talib.RSI(npClose, timeperiod=14) #相對強弱指數 (RSI)
        # slowk, slowd = talib.STOCH(npHigh, npLow, npClose, fastk_period=14, slowk_period=3, slowk_matype=0, slowd_period=3, slowd_matype=0) #隨機指標 (STOCH)
        # upperband, middleband, lowerband = talib.BBANDS(npClose, timeperiod=20, nbdevup=2, nbdevdn=2, matype=0) #布林帶 (BBANDS)
        # obv    = talib.OBV(npClose, npVolume) #平衡交易量 (OBV)
        # willr  = talib.WILLR(npHigh, npLow, npClose, timeperiod=14) #威廉指標 (WILLR)
        # atr    = talib.ATR(npHigh, npLow, npClose, timeperiod=14) #平均真實範圍 (ATR)
        # cci    = talib.CCI(npHigh, npLow, npClose, timeperiod=14) #商品通道指數 (CCI)
        # mom    = talib.MOM(npClose, timeperiod=10) #動量指標 (MOM)
        # sar    = talib.SAR(npHigh, npLow, acceleration=0.02, maximum=0.2) #龐氏指標 (SAR)
        # adx    = talib.ADX(npHigh, npLow, npClose, timeperiod=14) #平均方向性指數 (ADX)
        # trix   = talib.TRIX(npClose, timeperiod=30) #三重指數平滑異同移動平均線 (TRIX)
        # tema   = talib.TEMA(npClose, timeperiod=30) #三重平滑指標 (TEMA)
        # stddev = talib.STDDEV(npClose, timeperiod=5, nbdev=1) #股票或其他金融時間序列資料的標準差 (Standard Deviation)
        # varma  = talib.VAR(npClose, timeperiod=14, nbdev=1) #可變移動平均線 (VARMA)
        # print(F"平均方向性指數 (ADX) = {type(adx)} {adx}")

        # # 計算VMA (Volume Moving Average)
        # vma_short = talib.SMA(npVolume, timeperiod=iShortTerm)
        # vma_mid   = talib.SMA(npVolume, timeperiod=10)
        # vma_long  = talib.SMA(npVolume, timeperiod=20)

        # # 計算VROC (Volume Rate of Change)
        # vroc = talib.ROC(npVolume, timeperiod=12)

        # # 力量指標 (FI) 需要自訂計算
        # def calc_force_index(close, volume, period=13):
            # force = talib.ROC(close, 1) * volume
            # return talib.SMA(force, timeperiod=period)
        # force_index = calc_force_index(npClose, npVolume, period=13)

        # # 標準差通道 (Standard Deviation Channel)
        # upper_stddev, lower_stddev = middleband + stddev, middleband - stddev

        # # 其他指標計算
        # cmo  = talib.CMO(npClose, timeperiod=14) #Chande Momentum Oscillator (CMO)
        # emv  = talib.LINEARREG(npHigh - npLow, timeperiod=14) #Ease of Movement (EMV)
        # kama = talib.KAMA(npClose, timeperiod=30) #Kaufman Adaptive Moving Average (KAMA)
        # mfi  = talib.MFI(npHigh, npLow, npClose, npVolume, timeperiod=14) #Money Flow Index (MFI)
        # pvt  = talib.OBV(npClose, npVolume) #Price and Volume Trend (PVT)

        # # print(F"簡單移動平均線 (SMA) = {type(sma)} {sma}")
        # # print(F"指數移動平均線 (EMA) = {type(ema)} {ema}")
        # # print(F"移動平均收斂背離指標 (MACD) : macd = {type(macd)} {macd}")
        # # print(F"移動平均收斂背離指標 (MACD) : macdsignal = {type(macdsignal)} {macdsignal}")
        # # print(F"移動平均收斂背離指標 (MACD) : macdhist = {type(macdhist)} {macdhist}")
        # # print(F"相對強弱指數 (RSI) = {type(rsi)} {rsi}")
        # # print(F"隨機指標 (STOCH) : slowk = {type(slowk)} {slowk},slowd = {type(slowd)} {slowd}")



        # # 判斷 多頭 vs 空頭，並將資料儲存到 dictionary 
        # dictSignals = {            
            # "簡單移動平均線 (SMA)": "多頭" if npClose[-1] > sma[-1] else "空頭",
            # "指數移動平均線 (EMA)": "多頭" if npClose[-1] > ema[-1] else "空頭",
            # "移動平均收斂背離指標 (MACD)": "多頭" if macd[-1] > macdsignal[-1] else "空頭",
            # "相對強弱指數 (RSI)": "多頭" if rsi[-1] > 50 else "空頭",
            # "隨機指標 (STOCH)": "多頭" if slowk[-1] > slowd[-1] else "空頭",
            # "布林帶 (BBANDS)"     : "多頭" if npClose[-1] < upperband[-1] and npClose[-1] > lowerband[-1] else "空頭",
            # "平衡交易量 (OBV)"    : "多頭" if obv[-1] > obv[-2] else "空頭",
            # "威廉指標 (WILLR)"    : "多頭" if willr[-1] > -50 else "空頭",
            # "平均真實範圍 (ATR)"  : "多頭" if atr[-1] > atr[-2] else "空頭",
            # "商品通道指數 (CCI)"  : "多頭" if cci[-1] > 100 else "空頭",
            # "動量指標 (MOM)"      : "多頭" if mom[-1] > 0 else "空頭",
            # "龐氏指標 (SAR)"      : "多頭" if npClose[-1] > sar[-1] else "空頭",
            # "力量指標 (FORCE)"    : "多頭" if force_index[-1] > 0 else "空頭",
            # "標準差 (STDDEV)"     : "多頭" if npClose[-1] < upper_stddev[-1] and npClose[-1] > lower_stddev[-1] else "空頭",
            # "平均方向性指數 (ADX)": "多頭" if adx[-1] > 25 else "空頭",
            # "可變移動平均線 (VARMA)": "多頭" if npClose[-1] > varma[-1] else "空頭",
            # "成交量移動平均線 (VMA 短期)"     : "多頭" if npVolume[-1] > vma_short[-1] else "空頭",
            # "成交量移動平均線 (VMA 中期)"     : "多頭" if npVolume[-1] > vma_mid[-1] else "空頭",
            # "成交量移動平均線 (VMA 長期)"     : "多頭" if npVolume[-1] > vma_long[-1] else "空頭",
            # "成交量比率 (VROC)"               : "多頭" if vroc[-1] > 0 else "空頭",
            # "Chande Momentum Oscillator (CMO)": "多頭" if cmo[-1] > 0 else "空頭",
            # "Ease of Movement (EMV)"          : "多頭" if emv[-1] > 0 else "空頭",
            # "Kaufman Adaptive Moving Average (KAMA)": "多頭" if npClose[-1] > kama[-1] else "空頭",
            # "Money Flow Index (MFI)"          : "多頭" if mfi[-1] > 50 else "空頭",
            # "Price and Volume Trend (PVT)"    : "多頭" if pvt[-1] > pvt[-2] else "空頭",
            # }

        # print(F"\t {self.sHr1}")
        # print(F"\t dDate = {listDates[-1]}")
        # # print(F"npClose[-1] = {type(npClose[-1])} {npClose[-1]}")
        # # print(F"sma[-1] = {type(sma[-1])} {sma[-1]}")
        # # print(F"ema[-1] = {type(ema[-1])} {ema[-1]}")
        # # print(F"\t macd[-1] = {type(macd[-1])} {macd[-1]}")
        # # print(F"\t macdsignal[-1] = {type(macdsignal[-1])} {macdsignal[-1]}")
        # # print(F"\t rsi[-1] = {type(rsi[-1])} {rsi[-1]}")
        # # print(F"\t slowk[-1] = {type(slowk[-1])} {slowk[-1]}")
        # # print(F"\t slowd[-1] = {type(slowd[-1])} {slowd[-1]}")
        # print(F"\t adx[-1] = {type(adx[-1])} {adx[-1]}")
        # print('dictSignals = ',dictSignals)




    def __ExcelFormat(self,sFilePath):
        """
        調整Excel格式. (目前沒有使用)

        Args:
            sFilePath (str): 檔案路徑及檔案名稱 (ex:C:/Users/user/Desktop/股票技術指標輸出/2024-05-24_3703_欣陸_測試輸出.xlsx)
        Returns:
            N/A
        """
        # print('\t\t ===== ExcelFormat() ===== ')
        # print(F"\t\t sFilePath = {type(sFilePath)} {sFilePath}")
        
        # ------------------------------
        # 使用 openpyxl 調整格式
        # ------------------------------
        workbook = load_workbook(sFilePath)
        worksheet = workbook.active

        # 凍結頂端列
        worksheet.freeze_panes = "A2"

        # 設定字體為微軟正黑體
        font = Font(name='Microsoft JhengHei', size=10)
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = font

        # 設定「第01列」
        font = Font(name='Microsoft JhengHei', size=10, bold=True) # 設定 字體 = 微軟正黑體，粗體
        fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid') # 設定 單元格背景顏色
        # worksheet[1] 指的是第一列
        for cell in worksheet[1]:
            cell.alignment = Alignment(wrap_text=True)  # 自動換行
            cell.fill = fill  # 設定背景顏色
            cell.font = font  # 設定字體為微軟正黑體

        # 設定欄寬
        # worksheet.column_dimensions['A'].width = 3
        # worksheet.column_dimensions['B'].width = 13
        # worksheet.column_dimensions['C'].width = 3
        # columnsSize10 = ['D', 'E', 'F', 'G', 'H', 'I','J']
        # for col in columnsSize10:
        #     worksheet.column_dimensions[col].width = 10

        # 保存更改
        workbook.save(sFilePath)



    def verify(self):
        """
        回測程式.

        Args:
        Returns:
            datetime: 調整後的日期
        """
        print('\t ===== verify() ===== ')

        # ------------------------------
        # 組出要寫到 Excel 的內容
        # ------------------------------
        listOutput = []

        # Set 此次要迴測的股票
        stockList = ["3703:欣陸","2472:立隆電","4919:新唐","6191:精成科","6197:佳必琪","2353:宏碁","3532:台勝科","2451:創見","3017:奇鋐","2002:中鋼","3042:晶技"]
        # stockList = ["3703:欣陸"]

        iNo = 0
        for stock in stockList:
            iNo += iNo
            stockCode, stockName = stock.split(":")
            print(f"{iNo}.stockCode = {stockCode}, stockName = {stockName} {self.sHr1}")

            # ------------------------------
            # 讀取 Excel 檔案
            # ------------------------------
            sFilePath = f"C:/Users/user/Desktop/股票技術指標輸出/回測數據/{stockCode}_{stockName}_2021-01-04_to_2024-07-16_FromClass.xlsx"
            print(F"\t sFilePath = {sFilePath}")

            # 使用 pandas 讀取 Excel 檔案
            df = pd.read_excel(sFilePath, sheet_name='Sheet1')

            # # 迴圈遍歷 DataFrame 的每一列並印出每個欄位的值
            # # for index, row in df.iterrows():
            #     # print(f'Row {index}:')
            #     # for col in df.columns:
            #         # print(f'  {col}: {row[col]}')
            # # <class 'bool'>

            # ------------------------------
            # 算出每隻股票的買/賣日期及價格
            # ------------------------------
            # # listMyMoney = [] #<class 'list'>
            dictMoney = {}
            dateBuy = ""
            isBuy = False
            isSell = False
            i7Day = 0
            for i in range(2, len(df)):
                stockDate   = df.loc[i, '交易日期'] # <class 'str'>
                stockCode   = df.loc[i, '股票代號'] # <class 'numpy.int64'>
                stockName   = df.loc[i, '股票名稱'] # <class 'str'>
                fClosePrice = df.loc[i, '收盤價']   # <class 'numpy.float64'>
                iUpCount    = df.loc[i, '技術指標-多頭數量'] # <class 'numpy.int64'>
                iDownCount  = df.loc[i, '技術指標-空頭數量'] # <class 'numpy.int64'>
                iUpDays     = df.loc[i, '「多頭數量」連續增加天數'] # <class 'numpy.int64'>
                iDownDays   = df.loc[i, '「空頭數量」連續增加天數'] # <class 'numpy.int64'>
                # print(F"\t {i}.{stockDate} {self.sHr1}")
                # print(F"\t\t stockDate = ({type(stockDate)}) {stockDate}")
                # print(F"\t\t stockCode = ({type(stockCode)}) {stockCode}")
                # print(F"\t\t stockName = ({type(stockName)}) {stockName}")
                # print(F"\t\t fClosePrice = ({type(fClosePrice)}) {fClosePrice}")
                # print(F"\t\t iUpCount = ({type(iUpCount)}) {iUpCount}")
                # print(F"\t\t iDownCount = ({type(iDownCount)}) {iDownCount}")
                # print(F"\t\t iUpDays = ({type(iUpDays)}) {iUpDays}")
                # print(F"\t\t iDownDays = ({type(iDownDays)}) {iDownDays}")
            
                # ------------------------------
                # 條件1
                # 買:「多頭數量」連續增加天數 >= 3 天
                # 賣:「空頭數量」連續增加天數 >= 3 天
                # ------------------------------
                if iUpDays >=2 and not isBuy:
                    isBuy = True
                    isSell = False
                    
                    # 記錄 買的 日期 及 價格
                    dateBuy = stockDate
                    dictMoney[stockDate] = [stockDate, fClosePrice, iUpCount, iUpDays]
                    # print(F"\t\t [isBuy] ({type(isBuy)}) {isBuy}")

                if iDownDays >=3 and isBuy and not isSell:
                    isSell = True
                    isBuy = False
                    
                    # 如果日期已存在於 dictMoney ,記錄 賣的 日期 及 價格
                    if dateBuy in dictMoney:
                        dictMoney[dateBuy].extend([stockDate, fClosePrice, iDownCount, iDownDays])
                    # print(F"\t\t [isSell] ({type(isSell)}) {isSell}")

                # ------------------------------
                # 條件2
                # 買:「多頭數量」連續增加天數 >= 3 天
                # 賣:買進 5/7/14 天後就賣
                # ------------------------------
                # if iUpDays >=3 and not isBuy:
                #     isBuy = True
                #     isSell = False
                    
                #     # 記錄 買的 日期 及 價格
                #     dateBuy = stockDate
                #     dictMoney[stockDate] = [stockDate, fClosePrice, iUpCount, iUpDays]
                #     # print(F"\t\t [isBuy] ({type(isBuy)}) {isBuy}")
                # if isBuy:
                #     i7Day = i7Day+1
                # if i7Day >=14 and isBuy and not isSell:
                #     isSell = True
                #     isBuy = False
                #     i7Day = 0
                    
                #     # 如果日期已存在於 dictMoney ,記錄 賣的 日期 及 價格
                #     if dateBuy in dictMoney:
                #         dictMoney[dateBuy].extend([stockDate, fClosePrice, iDownCount, iDownDays])
                #     # print(F"\t\t [isSell] ({type(isSell)}) {isSell}")
        
            # 初始化總盈虧、總投資額和總報酬率
            total_gain_loss  = 0
            total_investment = 0
            total_roi_sum    = 0  # 總報酬率
            for key, value in dictMoney.items():
                valueSize = len(value)
                # print(F"\t valueSize = ({type(valueSize)}) {valueSize}")

                if valueSize != 8:
                    continue

                buy_date   = value[0] # 買入日期
                buy_price  = value[1] # 買入價格
                iUpCount   = value[2] # 買入日期-技術指標-多頭數量
                iUpDays    = value[3] # 買入日期-「多頭數量」連續增加天數'

                sell_date  = value[4] # 賣出日期
                sell_price = value[5] # 賣出日期
                iDownCount = value[6] # 賣出日期-技術指標-空頭數量
                iDownDays  = value[7] # 賣出日期-「空頭數量」連續增加天數
                print(F"\t {self.sHr2}")
                print(F"\t Key: {key},value: {value}")
                print(F"\t buy_date   = ({type(buy_date)}) {buy_date}")
                print(F"\t buy_price  = ({type(buy_price)}) {buy_price}")
                print(F"\t sell_date  = ({type(sell_date)}) {sell_date}")
                print(F"\t sell_price = ({type(sell_price)}) {sell_price}")

                # 計算每次買入1000股的盈虧
                gain_loss = (sell_price - buy_price) * 1000
                gain_loss = round(gain_loss, 2)
                print(F"\t gain_loss = ({type(gain_loss)}) {gain_loss}")
            
                # get 投資報酬率，並轉換為百分比
                roi = (sell_price - buy_price) / buy_price * 100
                roi = round(roi, 2)
                print(F"\t roi = ({type(roi)}) {roi}")

        #     # 累加總盈虧
        #     total_gain_loss += gain_loss
        #     print(F"\t total_gain_loss = ({type(total_gain_loss)}) {total_gain_loss}")
            
        #     # 累加總投資額
        #     total_investment += buy_price * 1000
        #     print(F"\t total_investment = ({type(total_investment)}) {total_investment}")
            
        #     # 累加每次的報酬率
        #     total_roi_sum += roi  # 加總每次的報酬率
        #     print(F"\t total_roi_sum = ({type(total_roi_sum)}) {total_roi_sum}")
                
                row = [stockCode,stockName,buy_date, buy_price, iUpCount, iUpDays, sell_date, sell_price, iDownCount, iDownDays, gain_loss, roi]
                listOutput.append(row)

        # ------------------------------
        # 將資料寫入 Excel
        # ------------------------------
        # Set 表頭
        # columns = ["股票代號", "股票名稱", "買進日期","買進價格", "賣出日期","賣出價格"]
        columns = ["股票代號","股票名稱","買進日期","買進價格", "買入日期-技術指標-多頭數量", "買入日期-「多頭數量」連續增加天數", "賣出日期","賣出價格","賣出日期-技術指標-空頭數量","賣出日期-「空頭數量」連續增加天數","每次買入1000股的盈虧","投資報酬率"]
        # print(F"columns type = {type(columns)}")

        df = pd.DataFrame(listOutput, columns=columns)

        # Set 檔案名稱
        sPath="C:/Users/user/Desktop/股票技術指標輸出"
        sFileName = f"2024-07-16_多頭連增2天買_空頭連增3天賣.xlsx"
        # sFileName = f"2024-07-16_多頭連增3天買_空頭連增3天賣.xlsx"
        # sFileName = f"2024-07-16_多頭連增3天買_買進7天賣.xlsx"
        # sFileName = f"2024-07-16_多頭連增3天買_買進5天賣.xlsx"
        # sFileName = f"2024-07-16_多頭連增3天買_買進14天賣.xlsx"

        # 構建完整的文件路徑
        sFilePath = f"{sPath}/{sFileName}"

        # 將資料保存到 Excel 文件
        df.to_excel(sFilePath, index=False)
        print(f"\t 1.資料已保存到 = {sFilePath}")

        self.__ExcelFormat(sFilePath)

